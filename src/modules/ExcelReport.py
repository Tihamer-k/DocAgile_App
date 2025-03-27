import os

from colorama import Fore, Style
from openpyxl.reader.excel import load_workbook

import src.utils.Settings as Settings
import pandas as pd
from pandas import ExcelWriter
from tabulate import tabulate


class ExcelReport:
    """ExcelReport Class

    This class represents a report generator for Excel files.

    Attributes:
        data_obj (pandas.DataFrame): The data object containing the report data.

    Methods:
        generate_excel_report(final_out_path: str) -> None:
            Generates an Excel report and saves it to the specified path.

    """
    def __init__(self):
        self.data_obj = None

    def generate_excel_report(self, final_out_path: str):
        if len(Settings.DIFF_DATA) > 0:
            self.data_obj = pd.DataFrame(Settings.DIFF_DATA)
            # df = pd.DataFrame(s.COMPONENT)
            try:
                # validar si existe o no el folder de salida:
                os.makedirs(os.path.dirname(final_out_path), exist_ok=True)

                writer = ExcelWriter(final_out_path, engine=None)
                # df.to_excel(writer, sheet_name="hoja1", index=False, startrow=2, startcol=2)
                self.data_obj.to_excel(writer, sheet_name="hoja1", index=False)
                writer.close()
                print(Fore.GREEN + "¡Data guardada!" + Style.RESET_ALL)
                modify_excel_report(final_out_path)
                try_open(final_out_path, self.data_obj)
            except (OSError, IOError) as e:
                print(Fore.RED + "Problema guardando la data: " + Style.RESET_ALL + e)


def try_open(final_out_path, data_obj):
    """
    Tries to open the Excel report at the specified path. If it fails, it prints the report in the terminal.

    :param final_out_path: The path of the Excel report to be opened.
    :param data_obj: pandas.DataFrame: The data object containing the report data.
    :return: None
    """
    try:
        print("Open Excel report: " + final_out_path)
        os.startfile(final_out_path)
    except Exception as e:
        print("Error opening Excel report: " + str(e))
        # Sin importar si se puede abrir o no el Excel creado mostramos el reporte en la terminal
        header_list = data_obj.columns
        print(tabulate(data_obj, headers=header_list, tablefmt="fancy_grid"))



def modify_excel_report(final_out_path: str):
    """
    Modifies the width of columns in an Excel report based on the content length of each cell.
    Also, it saves the modified report and opens it.

    :param final_out_path: The path of the Excel report to be modified.
    :return: None
    """
    try:
        wb = load_workbook(final_out_path)
        ws = wb['hoja1']
        for letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
            max_width = 0

            for row_num in range(1, ws.max_row + 1):
                if len(ws[f"{letter}{row_num}"].value) > max_width:
                    max_width = len(ws[f"{letter}{row_num}"].value)

            ws.column_dimensions[letter].width = max_width + 1
        wb.save(final_out_path)
        print("File saved at: " + final_out_path)
    except Exception as e:
        print("Error modifying excel report: " + str(e))
